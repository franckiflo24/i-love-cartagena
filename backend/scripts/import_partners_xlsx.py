"""
Import partners from /app/backend/data/BASE_DE_DATO_APP.xlsx into MongoDB.

Idempotent — re-running only inserts NEW partners (matched by normalised name).
"""
import os, re, asyncio, unicodedata
from datetime import datetime, timezone

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv

load_dotenv("/app/backend/.env")
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "test_database")
XLSX = os.getenv("PARTNERS_XLSX", "/app/backend/data/BASE_DE_DATO_APP.xlsx")

DEFAULT_COORDS = (10.4236, -75.5378)
DEFAULT_IMAGE = {
    "restaurant": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1200",
    "club":       "https://images.unsplash.com/photo-1571266028243-d220c6f3a8e9?w=1200",
    "wellness":   "https://images.unsplash.com/photo-1545205597-3d9d02c29597?w=1200",
    "beach_club": "https://images.unsplash.com/photo-1519046904884-53103b34b206?w=1200",
    "hotel":      "https://images.unsplash.com/photo-1542314831-068cd1dbfeeb?w=1200",
}

def _norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().lower()

def _instagram_handle(url):
    if not url: return ""
    url = str(url).strip()
    m = re.search(r"instagram\.com/([^/?]+)/?", url)
    return m.group(1) if m else ""

def _slug(s):
    s = _norm(s)
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s[:24]

def _subcat_slug(s):
    if not s: return ""
    raw = _norm(s).strip()
    mapping = {
        "del mar": "del_mar", "fast food": "fast_food",
        "vegetarianos/healty": "healthy", "vegetarianos / healty": "healthy",
        "unas": "unas", "uñas": "unas",
        "recuperacion": "recuperacion", "peluqueria": "peluqueria",
        "cafe": "cafe", "gastronomicos": "gastronomico",
        "italiano": "italiano", "arabe": "arabe", "internacional": "internacional",
        "colombiano": "colombiano", "mediterraneo": "mediterraneo",
        "spa": "spa", "fitness": "fitness", "sport": "sport", "yoga": "yoga",
    }
    if raw in mapping: return mapping[raw]
    return re.sub(r"[^a-z0-9]+", "_", raw).strip("_")

def parse_restaurantes(ws):
    out = []
    subcats = []
    for col in range(1, ws.max_column + 1, 4):
        v = ws.cell(row=3, column=col).value
        if v and isinstance(v, str) and v.strip() and "instagram" not in v.lower():
            subcats.append((col, v.strip()))
    for r in range(4, ws.max_row + 1):
        for name_col, label in subcats:
            name = ws.cell(row=r, column=name_col).value
            ig = ws.cell(row=r, column=name_col + 2).value
            if name and isinstance(name, str) and name.strip():
                out.append({"name": name.strip(), "category": "restaurant",
                            "subcategory": _subcat_slug(label),
                            "subcategory_label": label.strip(),
                            "instagram": _instagram_handle(ig) if ig else ""})
    return out

def parse_two_cols(ws, category, sub_slug, sub_label, layout=((1, 5), (7, 11))):
    out = []
    for r in range(3, ws.max_row + 1):
        for name_col, ig_col in layout:
            name = ws.cell(row=r, column=name_col).value
            ig = ws.cell(row=r, column=ig_col).value
            if name and isinstance(name, str) and name.strip():
                out.append({"name": name.strip(), "category": category,
                            "subcategory": sub_slug, "subcategory_label": sub_label,
                            "instagram": _instagram_handle(ig) if ig else ""})
    return out

def parse_wellness(ws):
    out = []
    subcats = []
    for col in range(1, ws.max_column + 1, 6):
        v = ws.cell(row=3, column=col).value
        if v and isinstance(v, str) and v.strip():
            subcats.append((col, v.strip()))
    for r in range(4, ws.max_row + 1):
        for name_col, label in subcats:
            name = ws.cell(row=r, column=name_col).value
            ig = ws.cell(row=r, column=name_col + 4).value
            if name and isinstance(name, str) and name.strip():
                out.append({"name": name.strip(), "category": "wellness",
                            "subcategory": _subcat_slug(label),
                            "subcategory_label": label.strip(),
                            "instagram": _instagram_handle(ig) if ig else ""})
    return out

def parse_hotel(ws):
    out = []
    for r in range(3, ws.max_row + 1):
        for name_col, ig_col in [(1, 2), (4, 5), (7, 8)]:
            name = ws.cell(row=r, column=name_col).value
            ig = ws.cell(row=r, column=ig_col).value
            if name and isinstance(name, str) and name.strip():
                out.append({"name": name.strip(), "category": "hotel",
                            "subcategory": "hotel_boutique",
                            "subcategory_label": "Hotel Boutique",
                            "instagram": _instagram_handle(ig) if ig else ""})
    return out

PARSERS = {
    "RESTAURANTES": parse_restaurantes,
    "DISCOTECAS Y BARES": lambda ws: parse_two_cols(ws, "club", "bar_disco", "Bar & Discoteca"),
    "WELLNESS": parse_wellness,
    "BEACH CLUB": lambda ws: parse_two_cols(ws, "beach_club", "beach_club", "Beach Club"),
    "HOTEL": parse_hotel,
}

async def main():
    print(f"Loading {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    all_rows = []
    for sheet, parser in PARSERS.items():
        if sheet in wb.sheetnames:
            rows = parser(wb[sheet])
            print(f"  {sheet}: {len(rows)} extracted")
            all_rows.extend(rows)
    seen = {}
    for r in all_rows:
        key = _norm(r["name"])
        if not key: continue
        if key in seen and not r.get("instagram"): continue
        if key in seen and seen[key].get("instagram") and not r.get("instagram"): continue
        seen[key] = r
    deduped = list(seen.values())
    print(f"Total: {len(all_rows)} raw, {len(deduped)} unique")

    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    existing = await db.partners.find({}, {"_id": 0, "partner_id": 1, "name": 1, "order": 1}).to_list(2000)
    existing_names = {_norm(p["name"]) for p in existing}
    max_order = max([p.get("order") or 0 for p in existing] + [0])
    print(f"Existing: {len(existing)} (max order={max_order})")
    to_insert = [r for r in deduped if _norm(r["name"]) not in existing_names]
    print(f"To insert: {len(to_insert)}")

    now = datetime.now(timezone.utc).isoformat()
    docs = []
    used_ids = {p["partner_id"] for p in existing if p.get("partner_id")}
    for i, r in enumerate(to_insert):
        slug = _slug(r["name"])
        pid = f"ptr_xl_{slug}_{i:03d}"
        n = 0
        while pid in used_ids:
            n += 1; pid = f"ptr_xl_{slug}_{i:03d}_{n}"
        used_ids.add(pid)
        cat_human = {"restaurant": "Restaurante", "club": "Bar & Discoteca",
                     "wellness": "Bienestar", "beach_club": "Beach Club", "hotel": "Hotel boutique"}[r["category"]]
        docs.append({
            "partner_id": pid, "name": r["name"].strip(),
            "description": f"{cat_human} en Cartagena — partner verificado de Amo Cartagena.",
            "category": r["category"], "subcategory": r["subcategory"],
            "subcategory_label": r["subcategory_label"],
            "image_url": DEFAULT_IMAGE[r["category"]],
            "location": {"lat": DEFAULT_COORDS[0], "lng": DEFAULT_COORDS[1]},
            "address": "", "booking_link": "", "price_range": "$$",
            "experience": "", "is_certified": True,
            "order": max_order + 100 + i, "tier": "popular",
            "instagram": r["instagram"] or "", "rating": 4.5, "reviews": 0,
            "membership_paid_until": None, "membership_status": "active",
            "membership_tier": "popular", "membership_plan": "free",
            "source": "xlsx_base_partners", "created_at": now,
        })
    if docs:
        await db.partners.insert_many(docs)
        print(f"INSERTED {len(docs)} ✅")
    else:
        print("Nothing to insert — DB already up-to-date.")

    print("\n=== Final state ===")
    async for row in db.partners.aggregate([{"$group": {"_id": "$category", "n": {"$sum": 1}}}, {"$sort": {"n": -1}}]):
        print(f"  {row['_id']}: {row['n']}")
    print(f"TOTAL: {await db.partners.count_documents({})}")

if __name__ == "__main__":
    asyncio.run(main())
